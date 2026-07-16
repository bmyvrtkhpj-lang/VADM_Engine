"""Parse the supplied Screener-style annual-fundamentals workbooks safely."""

from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook


def _find_row(ws, label: str, limit: int = 250) -> int:
    for row in range(1, min(ws.max_row, limit) + 1):
        value = ws.cell(row=row, column=1).value
        if value is not None and str(value).strip().casefold() == label.casefold():
            return row
    raise ValueError(f"Required row '{label}' was not found in the Data Sheet.")


def _row_values(ws, row: int) -> list:
    return [ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1)]


def extract_annual_fundamentals(workbook_path: str | Path) -> pd.DataFrame:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if "Data Sheet" not in workbook.sheetnames:
        raise ValueError("Workbook must contain a 'Data Sheet' worksheet.")
    sheet = workbook["Data Sheet"]
    pl_start, quarter_start = _find_row(sheet, "PROFIT & LOSS"), _find_row(sheet, "Quarters")
    report_row = profit_row = None
    for row in range(pl_start, quarter_start):
        label = sheet.cell(row=row, column=1).value
        label = str(label).strip() if label is not None else ""
        if label == "Report Date": report_row = report_row or row
        if label == "Net profit": profit_row = profit_row or row
    if report_row is None or profit_row is None:
        raise ValueError("Annual Report Date and Net profit rows are required.")
    derived_start = _find_row(sheet, "DERIVED:")
    shares_row = next((r for r in range(derived_start, min(derived_start + 12, sheet.max_row + 1))
                       if "Adjusted Equity Shares" in str(sheet.cell(r, 1).value)), None)
    if shares_row is None:
        raise ValueError("Adjusted Equity Shares row is required.")
    # Different downloaded sheets sometimes have unequal trailing blank/error cells.
    # Series construction aligns them without assuming perfectly matched row length.
    rows = pd.concat([
        pd.Series(_row_values(sheet, report_row), name="fiscal_year_end"),
        pd.Series(_row_values(sheet, profit_row), name="net_profit_cr"),
        pd.Series(_row_values(sheet, shares_row), name="shares_cr"),
    ], axis=1)
    rows["fiscal_year_end"] = pd.to_datetime(rows["fiscal_year_end"], errors="coerce")
    rows["net_profit_cr"] = pd.to_numeric(rows["net_profit_cr"], errors="coerce")
    rows["shares_cr"] = pd.to_numeric(rows["shares_cr"], errors="coerce")
    rows["annual_eps"] = rows["net_profit_cr"] / rows["shares_cr"].replace(0, pd.NA)
    rows = rows.dropna(subset=["fiscal_year_end", "annual_eps"]).sort_values("fiscal_year_end").drop_duplicates("fiscal_year_end")
    if rows.empty:
        raise ValueError("No valid annual EPS observations were found.")
    rows.attrs["company"] = sheet["B1"].value
    return rows.reset_index(drop=True)


def attach_point_in_time_pe(market: pd.DataFrame, annual: pd.DataFrame, filing_lag_days: int) -> pd.DataFrame:
    prices = market.reset_index().rename(columns={market.index.name or "index": "date"}).sort_values("date")
    filings = annual.assign(date=annual["fiscal_year_end"] + pd.Timedelta(days=filing_lag_days))[["date", "annual_eps"]].sort_values("date")
    merged = pd.merge_asof(prices, filings, on="date", direction="backward").set_index("date")
    merged["pe"] = merged["close"] / merged["annual_eps"]
    return merged
